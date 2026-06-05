#!/usr/bin/env python3
"""
Historical CyberVidya "Other Fees / Sanstha Collection" importer (DEV).

Parses the "Other Fees Sanstha Collection Report" workbook (one sheet per
COLLEGE code, a TALL row table) and creates one auto-submitted Journal Entry per
(college, fee-head, channel, account-head, date), booked into the college's
parent TRUST/SANSTHA company with immediate income recognition:

    Bank:  Dr {bank leaf in sanstha}  / Cr {fee-head income leaf in sanstha}
    Cash:  Dr {cash leaf in sanstha}  / Cr {fee-head income leaf in sanstha}

Sheet columns (by position 0..5): Payment Date | Account Group | Account Head |
Fees Head | Fees Head Short Name | Amount.  Account Group "Online Payment"=>bank,
"Cash"=>cash.  Fee heads are messy across colleges, so each is normalised and
folded to a CANONICAL head via SYNONYMS; one shared income leaf per canonical
head per sanstha (CLAUDE.md §14, locked decision: one ledger per head per sanstha).

Idempotency key on each JE: custom_cybervidya_ref
    HISTOF-{COLLEGE}-{CANONICAL_KEY}-{CASH-token|ACCT_LAST4}-{YYYYMMDD}

Two modes:
    --mode dryrun    Read-only. Resolves sansthas, prints the canonical-head plan,
                     unmapped labels, ledger plan, FY pre-flight, JE plan +
                     per-sheet reconciliation. NO writes.
    --mode execute   Creates the income group + leaves + bank/cash leaves +
                     CyberVidya Other Fees Mapping rows, then the JEs. ONLY after
                     RGI sign-off on the canonical-head plan.

Run on the dev box with the bench venv python, e.g.:
    cd ~/frappe-bench/sites
    ../env/bin/python ~/frappe-bench/apps/dux_cybervidya/dux_cybervidya/scripts/other_fees_import.py \
        --site erp.jewonline.in \
        --file /home/frappe/frappe-bench/sites/erp.jewonline.in/private/files/other_fees.xlsx \
        --mode dryrun
"""

import argparse
import os
import re
import sys
import traceback
from collections import defaultdict, OrderedDict
from decimal import Decimal, InvalidOperation


# ---------------------------------------------------------------------------
# Routing + canonical-head config (the parts RGI signs off in the dry run)
# ---------------------------------------------------------------------------

# College CV code -> exact ERPNext SANSTHA company (the "{ACRONYM} Society {City}"
# posting entities, confirmed against dev). None = PARKED (do not post).
# NOTE (RGI to ratify): GHRCEM -> Nagpur society, GHRCACS -> Pune society (it is the
# Pune-based "GHR CACS Pune"); GHRIBM/SRWC -> Jalgaon society. A dedicated entity
# "ASS For GHRCEM" also exists on dev — confirm GHRCEM's society before go-live.
COLLEGE_TO_SANSTHA = {
    "GHRCE":   "Ankush Shikshan Sanstha Society",
    "GHRIETN": "Ankush Shikshan Sanstha Society",
    "GHRILS":  "Ankush Shikshan Sanstha Society",
    "GHRCEM":  "GHREMF Society Nagpur",
    "GHRCACS": "GHREMF Society Pune",
    "GHRIBM":  "GHREF Society Jalgaon",
    "SRWC":    "GHREF Society Jalgaon",
    "GHRCCST": None,  # PARKED — sanstha unknown (RGI to confirm; CLAUDE.md §14 open item)
}

# Optional explicit override: logical sanstha name -> exact ERPNext Company name.
# Fill in after confirming names read-only on dev (Company.name / abbr) if the
# name-matching below is ambiguous or misses.
SANSTHA_COMPANY_OVERRIDE = {
    # "Ankush Shikshan Sanstha": "Ankush Shikshan Sanstha",
}

# Canonical head key -> display name used to build the sanstha income leaf
# ("{display} - {abbr}"). One leaf per key per sanstha.
CANONICAL_HEADS = {
    "EXAM":                "Examination Fees",
    "CONVOCATION":         "Convocation Fees",
    "PROSPECTUS":          "Prospectus Fees",
    "ALUMNI":              "Alumni Fund",
    "STUDENT_ACTIVITY":    "Student Activity Fees",
    "STUDENT_MEMBERSHIP":  "Student Membership Fees",
    "STUDENT_ENGAGEMENT":  "Student Engagement Fees",
    "TRAINING":            "Training Fees",
    "SKILL_DEV":           "Skill Development Fees",
    "PROF_DEV":            "Professional Development Fees",
    "PROJECT":             "Project Fees",
    "DEGREE":              "University Degree Fees",
    "PROVISIONAL_DEGREE":  "Provisional Degree Fees",
    "LABORATORY":          "Laboratory Fees",
    "LIBRARY_LAB":         "Library & Laboratory Fees",
    "SPORTS":              "University Sports Fees",
    "ESUVIDHA":            "E-Suvidha Fees",
    "PASSING_CERT":        "Passing Certificate Fees",
    "OTHER":               "Other Fees",
}

# Normalised-lower fee label -> canonical key. Folds case / spacing / plural /
# typos. Any label NOT here is reported as "unmapped" in the dry run (RGI fills
# the gap before execute). Seeded from the Apr-Jun 2026 workbook inventory.
SYNONYMS = {
    "exam fee": "EXAM", "exam fees": "EXAM",
    "examination fee": "EXAM", "examination fees": "EXAM",
    "convocation fee": "CONVOCATION", "convocation fees": "CONVOCATION",
    "prospectus fee": "PROSPECTUS", "prospectus fees": "PROSPECTUS",
    "propspectus fees": "PROSPECTUS", "propspectus fee": "PROSPECTUS",
    "alumni fund": "ALUMNI", "alumni fees": "ALUMNI", "alumni fee": "ALUMNI",
    "student activity": "STUDENT_ACTIVITY", "student activity fee": "STUDENT_ACTIVITY",
    "student activity fees": "STUDENT_ACTIVITY", "student activities": "STUDENT_ACTIVITY",
    "student membership fee": "STUDENT_MEMBERSHIP", "student membership fees": "STUDENT_MEMBERSHIP",
    "student engagement fee": "STUDENT_ENGAGEMENT", "student engagement fees": "STUDENT_ENGAGEMENT",
    "training fee": "TRAINING", "training fees": "TRAINING",
    "skill development fee": "SKILL_DEV", "skill development fees": "SKILL_DEV",
    "professional development fee": "PROF_DEV", "professional development fees": "PROF_DEV",
    "project fee": "PROJECT", "project fees": "PROJECT",
    "university degree fee": "DEGREE", "university degree fees": "DEGREE", "degree fee": "DEGREE",
    "provisional degree fee": "PROVISIONAL_DEGREE", "provisional degree fees": "PROVISIONAL_DEGREE",
    "laboratory fee": "LABORATORY", "laboratory fees": "LABORATORY",
    "library & laboratory": "LIBRARY_LAB", "library and laboratory": "LIBRARY_LAB",
    "library & laboratory fees": "LIBRARY_LAB",
    "university sports fee": "SPORTS", "university sports fees": "SPORTS",
    "e-suvidha fees": "ESUVIDHA", "e-suvidha fee": "ESUVIDHA", "esuvidha fees": "ESUVIDHA",
    "passing certificate fee": "PASSING_CERT", "passing certificate fees": "PASSING_CERT",
    "other fee": "OTHER", "other fees": "OTHER",
}

OTHER_FEES_GROUP = "CyberVidya Other Fees"      # group name (suffixed " - {abbr}")
OTHER_FEES_CASH_LEAF = "CyberVidya Other Fees Cash"   # shared cash leaf per sanstha

BANK_GROUPS = {"online payment", "bank transfer", "cheque", "demand draft", "online", "neft", "rtgs", "upi"}
CASH_GROUPS = {"cash"}

NBSP = "\xa0"
ACCT_RE = re.compile(r"(\d{6,})")
DATE_RE = re.compile(r"^\s*(\d{2})-(\d{2})-(\d{4})\s*$")


# ---------------------------------------------------------------------------
# Small parsing helpers (standalone — no frappe needed)
# ---------------------------------------------------------------------------
def clean_nom(s):
    if s is None:
        return ""
    s = str(s).replace(NBSP, " ")
    return re.sub(r"\s+", " ", s).strip()


def normalize_label(s):
    """Match key for fee labels: clean + lower-case (mirrors utils.normalize_fee_label)."""
    return clean_nom(s).lower()


def canonical_key(label_norm):
    return SYNONYMS.get(label_norm)


def extract_acct_no(cleaned_nom):
    matches = ACCT_RE.findall(cleaned_nom.replace(" ", ""))
    return matches[-1] if matches else None


def slug6(cleaned_nom):
    alnum = re.sub(r"[^A-Za-z0-9]", "", cleaned_nom).upper()
    return alnum[:6] or "UNKNWN"


def head_token(kind, account_head, acct_no):
    """Stable ref token for the channel/account-head."""
    if kind == "bank":
        return acct_no[-4:] if acct_no else slug6(account_head)
    s = re.sub(r"[^A-Za-z0-9]", "", account_head).upper()
    return s[:8] or "CASH"


def parse_date(cell):
    import datetime as _dt
    if isinstance(cell, (_dt.datetime, _dt.date)):
        return (cell.year, cell.month, cell.day)
    if cell is None:
        return None
    m = DATE_RE.match(str(cell))
    if not m:
        return None
    dd, mm, yyyy = m.groups()
    return (int(yyyy), int(mm), int(dd))


def to_decimal(cell):
    if cell is None or cell == "":
        return None
    try:
        d = Decimal(str(cell).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None
    return d if d > 0 else None


def dstr(d):
    return f"{d[0]:04d}-{d[1]:02d}-{d[2]:02d}"


def ddmy(d):
    return f"{d[0]:04d}{d[1]:02d}{d[2]:02d}"


def money(d):
    return f"{d:,.2f}"


# ---------------------------------------------------------------------------
# Workbook parsing (tall layout)
# ---------------------------------------------------------------------------
def parse_workbook_tall(path):
    """
    sheets = OrderedDict[sheet] -> {
        rows: [ {date,(y,m,d); channel 'bank'|'cash'; account_head; acct_no;
                 fee_raw; fee_norm; fee_short; amount Decimal} ],
        data_rows, min_date, max_date, total Decimal,
        anomalies: [(rownum, reason, detail)],
        labels: OrderedDict[fee_norm] -> {display, short, count, total},
        heads:  OrderedDict[(channel, account_head)] -> {acct_no, count, total},
    }
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    sheets = OrderedDict()

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows, anomalies = [], []
        dates_seen = set()
        total = Decimal(0)
        labels = OrderedDict()
        heads = OrderedDict()

        for r in range(2, ws.max_row + 1):   # row 1 is the header
            date_cell = ws.cell(row=r, column=1).value
            group = clean_nom(ws.cell(row=r, column=2).value)
            account_head = clean_nom(ws.cell(row=r, column=3).value)
            fee_raw = clean_nom(ws.cell(row=r, column=4).value)
            fee_short = clean_nom(ws.cell(row=r, column=5).value)
            amt = to_decimal(ws.cell(row=r, column=6).value)
            d = parse_date(date_cell)

            # silently skip fully-blank rows
            if d is None and not group and not fee_raw and amt is None and not account_head:
                continue
            if d is None:
                anomalies.append((r, "bad/blank date", repr(date_cell)))
                continue
            if not fee_raw:
                anomalies.append((r, "blank fee head", ""))
                continue
            if amt is None:
                anomalies.append((r, "blank/zero amount", ""))
                continue

            g = group.lower()
            if g in BANK_GROUPS:
                channel = "bank"
            elif g in CASH_GROUPS:
                channel = "cash"
            else:
                anomalies.append((r, f"unknown account group {group!r}", ""))
                continue

            acct_no = extract_acct_no(account_head) if channel == "bank" else None
            fee_norm = normalize_label(fee_raw)

            rows.append({
                "date": d, "channel": channel, "account_head": account_head,
                "acct_no": acct_no, "fee_raw": fee_raw, "fee_norm": fee_norm,
                "fee_short": fee_short, "amount": amt,
            })
            dates_seen.add(d)
            total += amt

            lab = labels.setdefault(fee_norm, {"display": fee_raw, "short": fee_short,
                                               "count": 0, "total": Decimal(0)})
            lab["count"] += 1
            lab["total"] += amt
            hk = (channel, account_head)
            h = heads.setdefault(hk, {"acct_no": acct_no, "count": 0, "total": Decimal(0)})
            h["count"] += 1
            h["total"] += amt

        sheets[sn] = {
            "rows": rows, "data_rows": len(rows),
            "min_date": min(dates_seen) if dates_seen else None,
            "max_date": max(dates_seen) if dates_seen else None,
            "total": total, "anomalies": anomalies,
            "labels": labels, "heads": heads,
        }
    return sheets


# ---------------------------------------------------------------------------
# Frappe-dependent resolution
# ---------------------------------------------------------------------------
class Resolver:
    def __init__(self, frappe):
        self.f = frappe

    def resolve_sanstha(self, sanstha_label):
        """logical sanstha name -> (company, abbr, method) or (None, None, reason)."""
        f = self.f
        if sanstha_label in SANSTHA_COMPANY_OVERRIDE:
            name = SANSTHA_COMPANY_OVERRIDE[sanstha_label]
            if f.db.exists("Company", name):
                return name, f.db.get_value("Company", name, "abbr"), "override"
            return None, None, f"override company {name!r} not found"
        if f.db.exists("Company", sanstha_label):
            return sanstha_label, f.db.get_value("Company", sanstha_label, "abbr"), "exact"
        rows = f.db.sql(
            "SELECT name, abbr FROM `tabCompany` WHERE UPPER(company_name)=UPPER(%s) OR UPPER(name)=UPPER(%s)",
            (sanstha_label, sanstha_label), as_dict=True)
        if len(rows) == 1:
            return rows[0].name, rows[0].abbr, "name-ci"
        if len(rows) > 1:
            return None, None, f"ambiguous name ({len(rows)} companies)"
        like = "%" + re.sub(r"[^A-Za-z0-9]+", "%", sanstha_label).strip("%") + "%"
        rows = f.db.sql(
            "SELECT name, abbr FROM `tabCompany` WHERE name LIKE %s OR company_name LIKE %s",
            (like, like), as_dict=True)
        if len(rows) == 1:
            return rows[0].name, rows[0].abbr, "fuzzy"
        if len(rows) > 1:
            return None, None, f"ambiguous fuzzy ({[r.name for r in rows][:5]})"
        return None, None, "no company match"

    def account_leaf_exists(self, name):
        rec = self.f.db.get_value("Account", name, ["is_group"], as_dict=True)
        if not rec:
            return False
        return not rec.is_group

    def indirect_income_group(self, company, abbr):
        f = self.f
        cand = f"Indirect Income - {abbr}"
        rec = f.db.get_value("Account", cand, ["name", "is_group", "root_type"], as_dict=True)
        if rec and rec.is_group and rec.root_type == "Income":
            return cand
        rows = f.db.sql("""
            SELECT name FROM `tabAccount`
            WHERE company=%s AND is_group=1 AND root_type='Income'
            ORDER BY lft LIMIT 1
        """, (company,), as_dict=True)
        return rows[0].name if rows else None

    def cash_parent(self, company, abbr):
        f = self.f
        rows = f.db.sql("""
            SELECT name FROM `tabAccount`
            WHERE company=%s AND is_group=1 AND account_type='Cash' ORDER BY lft LIMIT 1
        """, (company,), as_dict=True)
        if rows:
            return rows[0].name
        cand = f"Cash In Hand - {abbr}"
        rec = f.db.get_value("Account", cand, ["name", "is_group"], as_dict=True)
        if rec and rec.is_group:
            return cand
        rows = f.db.sql("""
            SELECT name FROM `tabAccount`
            WHERE company=%s AND is_group=1 AND root_type='Asset' AND name LIKE %s
            ORDER BY lft LIMIT 1
        """, (company, "%Cash%"), as_dict=True)
        return rows[0].name if rows else None

    def bank_parent(self, company, abbr):
        f = self.f
        cand = f"Bank Accounts - {abbr}"
        rec = f.db.get_value("Account", cand, ["name", "is_group"], as_dict=True)
        if rec and rec.is_group:
            return cand
        rows = f.db.sql("""
            SELECT name FROM `tabAccount`
            WHERE company=%s AND is_group=1 AND account_type='Bank' ORDER BY lft LIMIT 1
        """, (company,), as_dict=True)
        return rows[0].name if rows else None

    def find_bank_ledger(self, company, abbr, acct_no, account_head):
        """{action: matched|create|review, account, candidates, parent}."""
        f = self.f
        if acct_no:
            cands = f.db.sql("""
                SELECT name FROM `tabAccount`
                WHERE company=%s AND is_group=0 AND account_type='Bank' AND name LIKE %s
            """, (company, f"%{acct_no}%"), as_dict=True)
            names = [r.name for r in cands]
            if len(names) == 1:
                return {"action": "matched", "account": names[0], "candidates": names, "parent": None}
            if len(names) > 1:
                return {"action": "review", "account": None, "candidates": names, "parent": None}
        base = re.sub(r"[^A-Za-z0-9 -]", " ", account_head)
        base = re.sub(r"\s+", " ", base).strip() or "Bank"
        planned = f"{base} - {abbr}"
        return {"action": "create", "account": planned, "candidates": [],
                "parent": self.bank_parent(company, abbr), "account_name": base}

    def fiscal_year_covers(self, company, date_tuple):
        f = self.f
        ds = dstr(date_tuple)
        fys = f.db.sql("""
            SELECT name FROM `tabFiscal Year`
            WHERE disabled=0 AND year_start_date<=%s AND year_end_date>=%s
        """, (ds, ds), as_dict=True)
        for fy in fys:
            companies = f.db.sql(
                "SELECT company FROM `tabFiscal Year Company` WHERE parent=%s", (fy.name,), as_dict=True)
            if not companies:
                return True, fy.name
            if any(c.company == company for c in companies):
                return True, fy.name
        return False, None

    def je_for_ref(self, ref):
        f = self.f
        active = f.db.get_value("Journal Entry", {"custom_cybervidya_ref": ref, "docstatus": 1}, "name")
        if active:
            return "already_exists_active", active
        any_je = f.db.get_value("Journal Entry", {"custom_cybervidya_ref": ref}, "name")
        if any_je:
            return "already_exists_other", any_je
        cancelled = f.db.sql("""
            SELECT name FROM `tabJournal Entry`
            WHERE custom_cybervidya_ref LIKE %s AND docstatus=2 LIMIT 1
        """, (f"{ref}__CANCELLED__%",), as_dict=True)
        if cancelled:
            return "previously_cancelled", cancelled[0].name
        return "would_create", None


# ---------------------------------------------------------------------------
# Planning
# ---------------------------------------------------------------------------
def build_plan(sheets, resolver, exclude=None):
    """
    Returns dict with: resolved, parked, excluded, sanstha_plan, mapping_plan,
    je_plan, unmapped, reconciliation.
    """
    exclude = set(exclude or [])
    resolved = OrderedDict()       # sn -> {company, abbr, sanstha, method}
    parked = OrderedDict()         # sn -> {reason, rows, total, min, max}
    excluded = OrderedDict()
    # per-company account plan
    sanstha_plan = OrderedDict()   # company -> {abbr, income_group(planned), inc_parent,
                                   #             heads:set, needs_cash, cash_parent, cash_leaf,
                                   #             banks: {account_head -> res}}
    mapping_plan = OrderedDict()   # sn -> {sanstha, fee_heads:[(norm,disp,short,leaf)], channels:[(type,head,leaf)]}
    je_plan = []
    unmapped = OrderedDict()       # sn -> [(label_norm, display, short, total)]
    reconciliation = OrderedDict() # sn -> {parsed, planned, gap, unmapped_total, review_total}

    def sp_for(company, abbr):
        if company not in sanstha_plan:
            sanstha_plan[company] = {
                "abbr": abbr,
                "income_group": f"{OTHER_FEES_GROUP} - {abbr}",
                "inc_parent": resolver.indirect_income_group(company, abbr),
                "heads": set(), "needs_cash": False,
                "cash_parent": resolver.cash_parent(company, abbr),
                "cash_leaf": f"{OTHER_FEES_CASH_LEAF} - {abbr}",
                "banks": OrderedDict(),
                "errors": [],
            }
        return sanstha_plan[company]

    for sn, data in sheets.items():
        if sn in exclude:
            excluded[sn] = {"rows": data["data_rows"], "total": data["total"],
                            "min": data["min_date"], "max": data["max_date"]}
            continue
        label = COLLEGE_TO_SANSTHA.get(sn, "__UNROUTED__")
        if label is None:
            parked[sn] = {"reason": "sanstha unknown (parked in COLLEGE_TO_SANSTHA)",
                          "rows": data["data_rows"], "total": data["total"],
                          "min": data["min_date"], "max": data["max_date"]}
            continue
        if label == "__UNROUTED__":
            parked[sn] = {"reason": "college not in COLLEGE_TO_SANSTHA",
                          "rows": data["data_rows"], "total": data["total"],
                          "min": data["min_date"], "max": data["max_date"]}
            continue
        company, abbr, method = resolver.resolve_sanstha(label)
        if not company:
            parked[sn] = {"reason": f"sanstha {label!r}: {method}",
                          "rows": data["data_rows"], "total": data["total"],
                          "min": data["min_date"], "max": data["max_date"]}
            continue

        resolved[sn] = {"company": company, "abbr": abbr, "sanstha": label, "method": method}
        sp = sp_for(company, abbr)

        # --- canonical heads + income leaves ---
        head_leaf = {}   # canonical_key -> planned income leaf name
        sheet_unmapped = []
        unmapped_total = Decimal(0)
        for norm, meta in data["labels"].items():
            key = canonical_key(norm)
            if not key:
                sheet_unmapped.append((norm, meta["display"], meta["short"], meta["total"]))
                unmapped_total += meta["total"]
                continue
            sp["heads"].add(key)
            head_leaf[norm] = f"{CANONICAL_HEADS[key]} - {abbr}"
        if sheet_unmapped:
            unmapped[sn] = sheet_unmapped

        # --- channels (bank/cash ledgers in the sanstha) ---
        head_to_ledger = {}   # (channel, account_head) -> ledger name or None
        review_total = Decimal(0)
        for (channel, account_head), hmeta in data["heads"].items():
            if channel == "cash":
                sp["needs_cash"] = True
                head_to_ledger[(channel, account_head)] = sp["cash_leaf"]
            else:
                if account_head in sp["banks"]:
                    res = sp["banks"][account_head]
                else:
                    res = resolver.find_bank_ledger(company, abbr, hmeta["acct_no"], account_head)
                    sp["banks"][account_head] = res
                if res["action"] == "review":
                    head_to_ledger[(channel, account_head)] = None
                    review_total += hmeta_total(data, channel, account_head)
                else:
                    head_to_ledger[(channel, account_head)] = res["account"]

        # --- mapping rows for this college ---
        mp = {"sanstha": company, "fee_heads": [], "channels": []}
        for norm, meta in data["labels"].items():
            if norm in head_leaf:
                mp["fee_heads"].append((norm, meta["display"], meta["short"], head_leaf[norm]))
        for (channel, account_head), ledger in head_to_ledger.items():
            if ledger is not None:
                mp["channels"].append(("Bank" if channel == "bank" else "Cash",
                                       account_head, ledger))
        mapping_plan[sn] = mp

        # --- aggregate -> JE plan ---
        agg = defaultdict(lambda: Decimal(0))
        keymeta = {}
        for row in data["rows"]:
            key = canonical_key(row["fee_norm"])
            if not key:
                continue  # unmapped label -> excluded (counted in unmapped_total)
            ledger = head_to_ledger.get((row["channel"], row["account_head"]))
            if ledger is None:
                continue  # bank review/skip
            income_leaf = f"{CANONICAL_HEADS[key]} - {abbr}"
            tok = head_token(row["channel"], row["account_head"], row["acct_no"])
            ref = f"HISTOF-{sn}-{key}-{tok}-{ddmy(row['date'])}"
            agg_key = (row["date"], row["channel"], ledger, income_leaf, ref)
            agg[agg_key] += row["amount"]
            keymeta[agg_key] = {"channel": row["channel"], "key": key, "account_head": row["account_head"]}

        planned_total = Decimal(0)
        for (date, channel, ledger, income_leaf, ref), amount in agg.items():
            planned_total += amount
            je_plan.append({
                "sheet": sn, "company": company, "abbr": abbr, "date": date,
                "channel": channel, "ledger": ledger, "income_leaf": income_leaf,
                "amount": amount, "ref": ref, "key": keymeta[(date, channel, ledger, income_leaf, ref)]["key"],
                "account_head": keymeta[(date, channel, ledger, income_leaf, ref)]["account_head"],
            })

        reconciliation[sn] = {
            "parsed": data["total"], "planned": planned_total,
            "gap": data["total"] - planned_total,
            "unmapped_total": unmapped_total, "review_total": review_total,
        }

    je_plan.sort(key=lambda r: (r["sheet"], r["date"], r["income_leaf"], r["ledger"]))
    return {
        "resolved": resolved, "parked": parked, "excluded": excluded,
        "sanstha_plan": sanstha_plan, "mapping_plan": mapping_plan,
        "je_plan": je_plan, "unmapped": unmapped, "reconciliation": reconciliation,
    }


def hmeta_total(data, channel, account_head):
    return data["heads"][(channel, account_head)]["total"]


# ---------------------------------------------------------------------------
# Dry-run report
# ---------------------------------------------------------------------------
def build_dryrun_report(path, sheets, plan, resolver, stamp):
    resolved = plan["resolved"]; parked = plan["parked"]; excluded = plan["excluded"]
    sanstha_plan = plan["sanstha_plan"]; je_plan = plan["je_plan"]
    unmapped = plan["unmapped"]; recon = plan["reconciliation"]
    L = []
    def w(s=""):
        L.append(s)

    all_dates = [d for s in sheets.values() for d in (s["min_date"], s["max_date"]) if d]
    dmin = dstr(min(all_dates)) if all_dates else "—"
    dmax = dstr(max(all_dates)) if all_dates else "—"
    grand_parsed = sum((s["total"] for s in sheets.values()), Decimal(0))

    w("# CyberVidya Other-Fees Import — DRY RUN")
    w(f"_Generated {stamp} • mode=dryrun • NO database writes performed_")
    w("")
    w("## 1. Source summary")
    w("")
    w(f"- **File:** `{path}`")
    w(f"- **Sheets (colleges):** {len(sheets)}")
    w(f"- **Date range:** {dmin} → {dmax}")
    w(f"- **Grand total parsed (all sheets):** {money(grand_parsed)}")
    w("")

    w("## 2. College → sanstha resolution")
    w("")
    w("| Sheet | Sanstha (logical) | ERPNext company | Abbr | Via | Status |")
    w("|---|---|---|---|---|---|")
    for sn in sheets:
        if sn in resolved:
            r = resolved[sn]
            w(f"| {sn} | {r['sanstha']} | {r['company']} | {r['abbr']} | {r['method']} | RESOLVED |")
        elif sn in excluded:
            w(f"| {sn} | — | — | — | — | EXCLUDED (--exclude) |")
        else:
            p = parked[sn]
            w(f"| {sn} | — | — | — | — | **PARKED** ({p['reason']}) |")
    if parked:
        w("")
        w("### Parked sheets (NOT imported)")
        w("")
        w("| Sheet | Reason | Rows | Total |")
        w("|---|---|---|---|")
        for sn, p in parked.items():
            w(f"| {sn} | {p['reason']} | {p['rows']} | {money(p['total'])} |")
    w("")

    w("## 3. Canonical fee-head plan (per sanstha) — RGI SIGN-OFF")
    w("")
    w("Each sanstha gets one income leaf per canonical head, under a "
      "`CyberVidya Other Fees - {abbr}` group. Member-college labels fold into these:")
    w("")
    # group resolved sheets by company
    by_company = OrderedDict()
    for sn, r in resolved.items():
        by_company.setdefault(r["company"], []).append(sn)
    for company, sns in by_company.items():
        sp = sanstha_plan[company]
        w(f"### {company} ({sp['abbr']})")
        w(f"- Income group: `{sp['income_group']}`  (parent: "
          f"{('`'+sp['inc_parent']+'`') if sp['inc_parent'] else '**NO Indirect Income group — HARD ERROR**'})")
        w(f"- Income leaves to ensure ({len(sp['heads'])}):")
        for key in sorted(sp["heads"]):
            leaf = f"{CANONICAL_HEADS[key]} - {sp['abbr']}"
            exists = "exists" if resolver.account_leaf_exists(leaf) else "WILL CREATE"
            w(f"    - `{leaf}`  ({exists})")
        if sp["needs_cash"]:
            cp = sp["cash_parent"]
            w(f"- Cash leaf: `{sp['cash_leaf']}` "
              f"({'exists' if resolver.account_leaf_exists(sp['cash_leaf']) else 'WILL CREATE'}; "
              f"parent: {('`'+cp+'`') if cp else '**NO cash parent — HARD ERROR**'})")
        # label folding for member colleges
        w("- Label folding:")
        for sn in sns:
            for norm, meta in sheets[sn]["labels"].items():
                key = canonical_key(norm)
                tgt = f"{CANONICAL_HEADS[key]} - {sp['abbr']}" if key else "**UNMAPPED**"
                w(f"    - {sn}: `{meta['display']}` (`{meta['short']}`) → {tgt}")
        w("")

    if unmapped:
        w("## 3b. UNMAPPED labels (add to SYNONYMS before execute)")
        w("")
        w("| Sheet | Label (normalised) | Display | Short | Total |")
        w("|---|---|---|---|---|")
        for sn, items in unmapped.items():
            for norm, disp, short, tot in items:
                w(f"| {sn} | `{norm}` | {disp} | {short} | {money(tot)} |")
        w("")
    else:
        w("## 3b. UNMAPPED labels: none — every fee label folds to a canonical head ✔")
        w("")

    w("## 4. Bank / cash ledger plan (per sanstha)")
    w("")
    for company, sns in by_company.items():
        sp = sanstha_plan[company]
        w(f"### {company} ({sp['abbr']})")
        if sp["banks"]:
            for head, res in sp["banks"].items():
                if res["action"] == "matched":
                    w(f"- BANK `{head}` → matched `{res['account']}`")
                elif res["action"] == "create":
                    w(f"- BANK `{head}` → WILL CREATE `{res['account']}` (under "
                      f"{('`'+res['parent']+'`') if res['parent'] else '**NO bank parent — HARD ERROR**'})")
                elif res["action"] == "review":
                    w(f"- BANK `{head}` → **REVIEW** candidates {res['candidates']} (cells SKIPPED)")
        if sp["needs_cash"]:
            w(f"- CASH (all heads) → `{sp['cash_leaf']}`")
        w("")

    w("## 5. Fiscal-year pre-flight")
    w("")
    w("| Company | Sample date | FY covers? | Fiscal Year |")
    w("|---|---|---|---|")
    seen = set()
    for sn, r in resolved.items():
        comp = r["company"]
        if comp in seen:
            continue
        seen.add(comp)
        sample = sheets[sn]["min_date"] or (2026, 4, 1)
        ok, fy = resolver.fiscal_year_covers(comp, sample)
        w(f"| {comp} | {dstr(sample)} | {'YES' if ok else '**NO**'} | {fy or '—'} |")
    w("")

    w("## 6. Journal Entry plan + reconciliation")
    w("")
    total_amt = sum((j["amount"] for j in je_plan), Decimal(0))
    n_bank = sum(1 for j in je_plan if j["channel"] == "bank")
    n_cash = sum(1 for j in je_plan if j["channel"] == "cash")
    w(f"- **Total JEs planned:** {len(je_plan)}  ({n_bank} bank, {n_cash} cash)")
    w(f"- **Total amount planned:** {money(total_amt)}")
    w("")
    w("| Sheet | Parsed total | Planned (JE) total | Gap | Unmapped | Bank-review |")
    w("|---|---|---|---|---|---|")
    for sn, rc in recon.items():
        flag = "" if rc["gap"] == 0 else "  ⚠"
        w(f"| {sn} | {money(rc['parsed'])} | {money(rc['planned'])} | {money(rc['gap'])}{flag} "
          f"| {money(rc['unmapped_total'])} | {money(rc['review_total'])} |")
    w("")
    w("> Gap should be 0 for a clean run. Any gap = unmapped-label amount + bank-review-skipped amount.")
    w("")

    w("## 7. Idempotency pre-check")
    w("")
    buckets = defaultdict(int)
    for j in je_plan:
        status, _ = resolver.je_for_ref(j["ref"])
        buckets[status] += 1
    labels = {
        "would_create": "Will CREATE (ref not seen)",
        "already_exists_active": "Will SKIP (active JE already has this ref)",
        "already_exists_other": "Will SKIP (a draft/other JE has this ref)",
        "previously_cancelled": "Will CREATE fresh (only cancelled JEs carry this ref)",
    }
    w("| Outcome | Count |")
    w("|---|---|")
    for k in ("would_create", "previously_cancelled", "already_exists_active", "already_exists_other"):
        if buckets.get(k):
            w(f"| {labels[k]} | {buckets[k]} |")
    w("")

    anomalies = [(sn, s["anomalies"]) for sn, s in sheets.items() if s["anomalies"]]
    if anomalies:
        w("## 8. Skipped rows (anomalies)")
        w("")
        w("| Sheet | Row | Reason | Detail |")
        w("|---|---|---|---|")
        for sn, items in anomalies:
            for rn, reason, detail in items[:25]:
                w(f"| {sn} | {rn} | {reason} | {detail} |")
        w("")

    w("## 9. First 20 planned JEs (sanity sample)")
    w("")
    w("| Sheet | Date | Ch | Dr ledger | Cr income leaf | Amount | Ref |")
    w("|---|---|---|---|---|---|---|")
    for j in je_plan[:20]:
        w(f"| {j['sheet']} | {dstr(j['date'])} | {j['channel']} | {j['ledger']} "
          f"| {j['income_leaf']} | {money(j['amount'])} | {j['ref']} |")
    w("")
    w("---")
    w("")
    w("**Dry run complete. Resolve UNMAPPED labels + confirm the canonical plan + "
      "any HARD ERRORs, then re-run with --mode execute.**")
    return "\n".join(L)


# ---------------------------------------------------------------------------
# Execute (guarded)
# ---------------------------------------------------------------------------
def _default_cc(frappe, company):
    cc = frappe.get_cached_value("Company", company, "cost_center")
    if cc:
        return cc
    abbr = frappe.get_cached_value("Company", company, "abbr")
    return (frappe.db.get_value("Cost Center", f"Main - {abbr}", "name")
            or frappe.db.get_value("Cost Center", {"company": company, "is_group": 0}, "name"))


def _ensure_account(frappe, account_name, company, parent, abbr, is_group=0, account_type=None):
    full = f"{account_name} - {abbr}"
    if frappe.db.exists("Account", full):
        return full, False
    acc = frappe.new_doc("Account")
    acc.account_name = account_name
    acc.company = company
    acc.parent_account = parent
    acc.is_group = is_group
    if account_type:
        acc.account_type = account_type
    acc.account_currency = frappe.db.get_value("Company", company, "default_currency")
    acc.insert(ignore_permissions=True)
    return acc.name, True


def execute(frappe, sheets, plan, resolver, out_dir, stamp):
    frappe.set_user("Administrator")
    resolved = plan["resolved"]; sanstha_plan = plan["sanstha_plan"]
    mapping_plan = plan["mapping_plan"]; je_plan = plan["je_plan"]

    created_accounts, created_map_parents, created_map_rows = [], [], []
    errors = []

    # --- 5a: per-sanstha accounts (income group + leaves + bank/cash) ---
    for company, sp in sanstha_plan.items():
        abbr = sp["abbr"]
        try:
            if not sp["inc_parent"]:
                sp["errors"].append("No Indirect Income group; income leaves NOT created")
                errors.append({"phase": "income-group", "company": company,
                               "error": "no Indirect Income group", "tb": ""})
                continue
            grp, made = _ensure_account(frappe, OTHER_FEES_GROUP, company, sp["inc_parent"], abbr, is_group=1)
            if made:
                created_accounts.append(grp)
            for key in sorted(sp["heads"]):
                leaf, made = _ensure_account(frappe, CANONICAL_HEADS[key], company, grp, abbr, is_group=0)
                if made:
                    created_accounts.append(leaf)
            if sp["needs_cash"]:
                if not sp["cash_parent"]:
                    sp["errors"].append("No cash parent; cash leaf NOT created")
                    errors.append({"phase": "cash", "company": company,
                                   "error": "no cash parent", "tb": ""})
                else:
                    leaf, made = _ensure_account(frappe, OTHER_FEES_CASH_LEAF, company,
                                                 sp["cash_parent"], abbr, is_group=0, account_type="Cash")
                    if made:
                        created_accounts.append(leaf)
            for head, res in sp["banks"].items():
                if res["action"] == "create" and res["parent"]:
                    leaf, made = _ensure_account(frappe, res["account_name"], company,
                                                 res["parent"], abbr, is_group=0, account_type="Bank")
                    if made:
                        created_accounts.append(leaf)
            frappe.db.commit()
        except Exception as e:
            frappe.db.rollback()
            errors.append({"phase": "accounts", "company": company,
                           "error": f"{type(e).__name__}: {e}", "tb": traceback.format_exc()})

    # --- 5b: mapping parents + child rows ---
    for sn, mp in mapping_plan.items():
        try:
            if not frappe.db.exists("CyberVidya Other Fees Mapping", sn):
                doc = frappe.new_doc("CyberVidya Other Fees Mapping")
                doc.cybervidya_institution = sn
                doc.sanstha_company = mp["sanstha"]
                doc.insert(ignore_permissions=True)
                created_map_parents.append(sn)
            doc = frappe.get_doc("CyberVidya Other Fees Mapping", sn)
            existing_labels = {r.fee_label for r in doc.fee_heads}
            existing_ch = {(r.channel_type, r.cybervidya_account_head) for r in doc.channels}
            changed = False
            for norm, disp, short, leaf in mp["fee_heads"]:
                if norm not in existing_labels and frappe.db.exists("Account", leaf):
                    doc.append("fee_heads", {"fee_label": norm, "fee_label_display": disp,
                                             "fee_short": short, "income_account": leaf})
                    changed = True
                    created_map_rows.append((sn, "fee", norm, leaf))
            for ctype, head, leaf in mp["channels"]:
                if (ctype, clean_nom(head)) not in existing_ch and frappe.db.exists("Account", leaf):
                    doc.append("channels", {"channel_type": ctype,
                                            "cybervidya_account_head": head, "ledger_account": leaf})
                    changed = True
                    created_map_rows.append((sn, "channel", f"{ctype}:{head}", leaf))
            if changed:
                doc.save(ignore_permissions=True)
            frappe.db.commit()
        except Exception as e:
            frappe.db.rollback()
            errors.append({"phase": "mapping", "sheet": sn,
                           "error": f"{type(e).__name__}: {e}", "tb": traceback.format_exc()})

    # --- 5c: journal entries ---
    created, skipped, failed = [], [], []
    n = 0
    for j in je_plan:
        n += 1
        ref = j["ref"]
        try:
            status, existing = resolver.je_for_ref(ref)
            if status in ("already_exists_active", "already_exists_other"):
                skipped.append((ref, existing))
                continue
            # accounts must exist (created in 5a); skip+record if not
            if not frappe.db.exists("Account", j["ledger"]) or not frappe.db.exists("Account", j["income_leaf"]):
                failed.append({"ref": ref, "error": "ledger or income leaf missing (account-creation blocked)", "tb": ""})
                continue
            je = frappe.new_doc("Journal Entry")
            je.voucher_type = "Journal Entry"
            je.company = j["company"]
            je.posting_date = frappe.utils.getdate(dstr(j["date"]))
            je.user_remark = (f"Historical CyberVidya other-fee {CANONICAL_HEADS[j['key']]} "
                              f"[{j['sheet']}] {j['channel']} {ref}")
            je.custom_cybervidya_ref = ref
            amt = float(j["amount"])
            cc = _default_cc(frappe, j["company"])

            def line(account, dr, cr):
                row = {"account": account, "debit_in_account_currency": dr,
                       "credit_in_account_currency": cr}
                root = frappe.get_cached_value("Account", account, "root_type")
                if root in ("Income", "Expense"):
                    row["cost_center"] = cc
                return row

            je.append("accounts", line(j["ledger"], amt, 0))         # Dr bank/cash
            je.append("accounts", line(j["income_leaf"], 0, amt))    # Cr income head
            je.insert(ignore_permissions=True)
            je.submit()
            frappe.db.commit()
            created.append((ref, je.name))
        except (frappe.UniqueValidationError, frappe.DuplicateEntryError):
            frappe.db.rollback()
            existing = frappe.db.get_value("Journal Entry", {"custom_cybervidya_ref": ref}, "name")
            skipped.append((ref, existing))
        except Exception as e:
            frappe.db.rollback()
            failed.append({"ref": ref, "error": f"{type(e).__name__}: {e}", "tb": traceback.format_exc()})
        if n % 25 == 0:
            print(f"  ... {n}/{len(je_plan)} ({len(created)} created, {len(skipped)} skipped, {len(failed)} failed)")

    # --- report ---
    L = ["# CyberVidya Other-Fees Import — EXECUTION RUN",
         f"_Generated {stamp} • mode=execute_", "",
         f"- JEs created: **{len(created)}**",
         f"- JEs skipped (already existed): {len(skipped)}",
         f"- JEs failed: {len(failed)}",
         f"- Accounts created: {len(created_accounts)}",
         f"- Mapping parents created: {len(created_map_parents)}",
         f"- Mapping child rows added: {len(created_map_rows)}", ""]
    if created_accounts:
        L.append("## Accounts created")
        L += [f"- {a}" for a in created_accounts]
        L.append("")
    by_co = defaultdict(lambda: Decimal(0))
    ref_to_j = {j["ref"]: j for j in je_plan}
    for ref, _ in created:
        j = ref_to_j.get(ref)
        if j:
            by_co[j["company"]] += j["amount"]
    L.append("## Posted totals per sanstha (created JEs only)")
    L.append("")
    L.append("| Company | Amount posted |")
    L.append("|---|---|")
    for co, amt in sorted(by_co.items()):
        L.append(f"| {co} | {money(amt)} |")
    L.append("")
    if failed:
        L.append("## Failures")
        L += [f"- `{f['ref']}`: {f['error']}" for f in failed]
        L.append("")

    report = "\n".join(L)
    rpath = os.path.join(out_dir, f"other_fees_import_run_{stamp}.md")
    with open(rpath, "w") as fh:
        fh.write(report)
    print(report)
    print(f"\n[run report written to {rpath}]")
    if errors or failed:
        epath = os.path.join(out_dir, f"other_fees_import_errors_{stamp}.log")
        with open(epath, "w") as fh:
            for e in errors:
                fh.write(f"=== {e.get('phase')} {e.get('company') or e.get('sheet')} ===\n{e.get('tb','')}\n{e.get('error','')}\n")
            for f in failed:
                fh.write(f"=== JE {f['ref']} ===\n{f.get('tb','')}\n{f.get('error','')}\n")
        print(f"[error log written to {epath}]")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--site", required=True)
    ap.add_argument("--file", required=True)
    ap.add_argument("--mode", choices=["dryrun", "execute"], default="dryrun")
    ap.add_argument("--out-dir", default="/tmp")
    ap.add_argument("--stamp", default="run")
    ap.add_argument("--exclude", default="", help="comma-separated sheet names to skip entirely")
    args = ap.parse_args()
    exclude = [s.strip() for s in args.exclude.split(",") if s.strip()]

    if not os.path.exists(args.file):
        print(f"ERROR: file not found: {args.file}", file=sys.stderr)
        sys.exit(2)

    print(f"[reading workbook] {args.file}")
    sheets = parse_workbook_tall(args.file)
    print(f"[parsed] {len(sheets)} sheets")

    os.chdir("/home/frappe/frappe-bench/sites")
    import frappe
    frappe.init(site=args.site)
    frappe.connect()
    try:
        resolver = Resolver(frappe)
        if exclude:
            print(f"[exclude] skipping sheets: {exclude}")
        plan = build_plan(sheets, resolver, exclude=exclude)

        if args.mode == "dryrun":
            report = build_dryrun_report(args.file, sheets, plan, resolver, args.stamp)
            rpath = os.path.join(args.out_dir, f"other_fees_import_dryrun_{args.stamp}.md")
            with open(rpath, "w") as fh:
                fh.write(report)
            print(report)
            print(f"\n[dry-run report written to {rpath}]")
        else:
            print("[MODE=execute] creating accounts, mappings, and JEs...")
            execute(frappe, sheets, plan, resolver, args.out_dir, args.stamp)
    finally:
        frappe.destroy()


if __name__ == "__main__":
    main()
