#!/usr/bin/env python3
"""
Historical CyberVidya fee-collection importer (DEV).

Parses a "Daily Fees Collection Report" workbook (one sheet per CyberVidya
institution code, a date x channel-nomenclature matrix) and creates one
auto-submitted Journal Entry per (company, target-ledger, date), matching the
live API's accounting model:

    Bank channel: Dr {bank leaf ledger}        / Cr Student Receivable Cybervidya - {ABBR}
    Cash channel: Dr Cash Cyber Vidhya - {ABBR} / Cr Student Receivable Cybervidya - {ABBR}

Idempotency key on each JE: custom_cybervidya_ref
    Cash: HIST-{CV_CODE}-CASH-{YYYYMMDD}
    Bank: HIST-{CV_CODE}-{ACCT_LAST4}-{YYYYMMDD}  (slug fallback if no digits)

Two modes:
    --mode dryrun    Phases 1-4. Read-only. Produces a dry-run report. No writes.
    --mode execute   Phase 5. Creates accounts/mappings then JEs. ONLY run after approval.

Run on the dev box with the bench venv python, e.g.:
    cd ~/frappe-bench/sites
    ../env/bin/python ~/frappe-bench/apps/dux_cybervidya/dux_cybervidya/scripts/historical_import.py \
        --site erp.jewonline.in \
        --file /home/frappe/frappe-bench/sites/erp.jewonline.in/private/files/hist.xlsx \
        --mode dryrun
"""

import argparse
import os
import re
import sys
import traceback
from collections import defaultdict, OrderedDict
from decimal import Decimal, InvalidOperation


# ---------------------------------------------------------------------------
# Channel group classification (driven by row-1 merged header)
# ---------------------------------------------------------------------------
BANK_GROUPS = {"Bank Transfer", "Online Payment", "Cheque", "Demand Draft"}
CASH_GROUPS = {"Cash"}
SKIP_GROUPS = {"Jv Institute Others"}
DATE_GROUP = "Payment Date"

NBSP = "\xa0"
ACCT_RE = re.compile(r"(\d{6,})")          # last run of >=6 digits
DATE_RE = re.compile(r"^\s*(\d{2})-(\d{2})-(\d{4})\s*$")


# ---------------------------------------------------------------------------
# Workbook parsing (no frappe needed)
# ---------------------------------------------------------------------------
def clean_nom(s):
    """Normalise a nomenclature string: NBSP->space, collapse ws, strip."""
    if s is None:
        return ""
    s = str(s).replace(NBSP, " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def extract_acct_no(cleaned_nom):
    """Last run of >=6 digits, or None."""
    matches = ACCT_RE.findall(cleaned_nom.replace(" ", ""))
    return matches[-1] if matches else None


def slug6(cleaned_nom):
    """First 6 alphanumeric chars, uppercased (fallback ref token)."""
    alnum = re.sub(r"[^A-Za-z0-9]", "", cleaned_nom).upper()
    return alnum[:6] or "UNKNWN"


def parse_date(cell):
    """DD-MM-YYYY string (or a datetime/date) -> (yyyy, mm, dd) tuple or None."""
    import datetime as _dt
    if isinstance(cell, (_dt.datetime, _dt.date)):
        return (cell.year, cell.month, cell.day)
    if cell is None:
        return None
    m = DATE_RE.match(str(cell))
    if not m:
        return None
    dd, mm, yyyy = m.groups()
    return (int(yyyy), int(mm), int(dd))


def to_decimal(cell):
    """Coerce a cell to Decimal; return None if empty/zero/invalid."""
    if cell is None or cell == "":
        return None
    try:
        d = Decimal(str(cell).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None
    return d if d > 0 else None


def carry_forward_row1(ws):
    """Return per-column row-1 group label, carrying merged headers forward."""
    groups, last = [], None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v not in (None, ""):
            last = str(v).strip()
        groups.append(last)
    return groups


def parse_workbook(path):
    """
    Parse the workbook into a structure:
      sheets = OrderedDict[sheet_name] -> {
        'columns': [ {col, group, kind('bank'|'cash'|'skip'|'empty'),
                      raw_nom, clean_nom, acct_no} ],
        'cells': list of {date:(y,m,d), col, amount:Decimal, kind, clean_nom, acct_no, raw_nom},
        'data_rows': int, 'min_date', 'max_date',
        'skipped_jv': {'rows': int, 'amount': Decimal},
        'empty_cols': [col,...],
      }
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    sheets = OrderedDict()

    for sn in wb.sheetnames:
        ws = wb[sn]
        groups = carry_forward_row1(ws)

        columns = []
        for c in range(1, ws.max_column + 1):
            group = groups[c - 1]
            if c == 1 or group == DATE_GROUP:
                continue
            raw = ws.cell(row=2, column=c).value
            cnom = clean_nom(raw)
            if group in SKIP_GROUPS:
                kind = "skip"
            elif group in CASH_GROUPS:
                kind = "cash"
            elif group in BANK_GROUPS:
                kind = "bank"
            else:
                kind = "unknown"
            if not cnom and kind in ("bank", "cash"):
                kind = "empty"
            columns.append({
                "col": c, "group": group, "kind": kind,
                "raw_nom": raw, "clean_nom": cnom,
                "acct_no": extract_acct_no(cnom) if kind == "bank" else None,
            })

        cells = []
        skipped_jv_rows = 0
        skipped_jv_amount = Decimal(0)
        dates_seen = set()
        empty_cols = [col["col"] for col in columns if col["kind"] == "empty"]

        for r in range(3, ws.max_row + 1):
            d = parse_date(ws.cell(row=r, column=1).value)
            if d is None:
                continue
            row_had_jv = False
            for col in columns:
                amt = to_decimal(ws.cell(row=r, column=col["col"]).value)
                if amt is None:
                    continue
                if col["kind"] == "skip":
                    skipped_jv_amount += amt
                    row_had_jv = True
                    continue
                if col["kind"] in ("empty", "unknown"):
                    continue
                dates_seen.add(d)
                cells.append({
                    "date": d, "col": col["col"], "amount": amt,
                    "kind": col["kind"], "clean_nom": col["clean_nom"],
                    "acct_no": col["acct_no"], "raw_nom": col["raw_nom"],
                    "group": col["group"],
                })
            if row_had_jv:
                skipped_jv_rows += 1

        sheets[sn] = {
            "columns": columns,
            "cells": cells,
            "data_rows": sum(1 for r in range(3, ws.max_row + 1)
                             if parse_date(ws.cell(row=r, column=1).value)),
            "min_date": min(dates_seen) if dates_seen else None,
            "max_date": max(dates_seen) if dates_seen else None,
            "skipped_jv": {"rows": skipped_jv_rows, "amount": skipped_jv_amount},
            "empty_cols": empty_cols,
        }
    return sheets


# ---------------------------------------------------------------------------
# Frappe-dependent resolution (Phases 1-4 read-only)
# ---------------------------------------------------------------------------
def dstr(d):
    return f"{d[0]:04d}-{d[1]:02d}-{d[2]:02d}"


def ddmy(d):
    return f"{d[0]:04d}{d[1]:02d}{d[2]:02d}"


class Resolver:
    def __init__(self, frappe):
        self.f = frappe

    def resolve_company(self, sheet_name):
        """(company, abbr, method) or (None, None, reason)."""
        f = self.f
        company = f.db.get_value(
            "CyberVidya Account Mapping",
            {"cybervidya_institution": sheet_name}, "company")
        if company:
            abbr = f.db.get_value("Company", company, "abbr")
            return company, abbr, "mapping"
        # fallback: abbr match (case-insensitive)
        rows = f.db.sql(
            "SELECT name, abbr FROM `tabCompany` WHERE UPPER(abbr)=UPPER(%s)",
            (sheet_name,), as_dict=True)
        if len(rows) == 1:
            return rows[0].name, rows[0].abbr, "abbr"
        if len(rows) > 1:
            return None, None, f"ambiguous abbr ({len(rows)} companies)"
        return None, None, "no mapping, no abbr match"

    def head_exists_leaf(self, name):
        rec = self.f.db.get_value("Account", name, ["is_group"], as_dict=True)
        if not rec:
            return False, "missing"
        if rec.is_group:
            return False, "is_group"
        return True, "ok"

    def find_bank_ledger(self, company, acct_no, raw_nom, sheet_name):
        """
        Returns dict: {action: 'matched'|'maprow'|'create'|'review',
                       account: str|None, candidates: [...], parent: str|None}
        """
        f = self.f
        # 4) existing CyberVidya Bank Map row (verbatim nomenclature)?
        maprow = f.db.sql("""
            SELECT bm.bank_account
            FROM `tabCyberVidya Bank Map` bm
            JOIN `tabCyberVidya Account Mapping` am ON bm.parent = am.name
            WHERE am.cybervidya_institution = %s AND bm.cybervidya_bank = %s
            LIMIT 1
        """, (sheet_name, raw_nom), as_dict=True)
        if maprow:
            return {"action": "maprow", "account": maprow[0].bank_account,
                    "candidates": [], "parent": None}
        # 2) acct_no LIKE match among Bank-type leaves
        if acct_no:
            cands = f.db.sql("""
                SELECT name FROM `tabAccount`
                WHERE company=%s AND is_group=0 AND account_type='Bank'
                  AND name LIKE %s
            """, (company, f"%{acct_no}%"), as_dict=True)
            names = [r.name for r in cands]
            if len(names) == 1:
                return {"action": "matched", "account": names[0],
                        "candidates": names, "parent": None}
            if len(names) > 1:
                return {"action": "review", "account": None,
                        "candidates": names, "parent": None}
        # 3) zero matches -> plan creation
        parent = self._bank_parent(company)
        return {"action": "create", "account": None, "candidates": [],
                "parent": parent}

    def _bank_parent(self, company):
        f = self.f
        abbr = f.db.get_value("Company", company, "abbr")
        cand = f"Bank Accounts - {abbr}"
        rec = f.db.get_value("Account", cand, ["name", "is_group"], as_dict=True)
        if rec and rec.is_group:
            return cand
        rows = f.db.sql("""
            SELECT name FROM `tabAccount`
            WHERE company=%s AND is_group=1 AND account_type='Bank'
            ORDER BY lft LIMIT 1
        """, (company,), as_dict=True)
        return rows[0].name if rows else None

    def planned_account_name(self, company, clean_nom):
        abbr = self.f.db.get_value("Company", company, "abbr")
        base = re.sub(r"[^A-Za-z0-9 -]", " ", clean_nom)
        base = re.sub(r"\s+", " ", base).strip()
        return f"{base} - {abbr}"

    def fiscal_year_covers(self, company, date_tuple):
        """Is there an enabled Fiscal Year covering this date for this company?"""
        f = self.f
        ds = dstr(date_tuple)
        fys = f.db.sql("""
            SELECT name FROM `tabFiscal Year`
            WHERE disabled=0 AND year_start_date<=%s AND year_end_date>=%s
        """, (ds, ds), as_dict=True)
        for fy in fys:
            companies = f.db.sql("""
                SELECT company FROM `tabFiscal Year Company` WHERE parent=%s
            """, (fy.name,), as_dict=True)
            if not companies:
                return True, fy.name  # global FY (no company restriction)
            if any(c.company == company for c in companies):
                return True, fy.name
        return False, None

    def je_for_ref(self, ref):
        """Return (status, je_name) for idempotency precheck."""
        f = self.f
        active = f.db.get_value(
            "Journal Entry", {"custom_cybervidya_ref": ref, "docstatus": 1}, "name")
        if active:
            return "already_exists_active", active
        any_je = f.db.get_value(
            "Journal Entry", {"custom_cybervidya_ref": ref}, "name")
        if any_je:
            return "already_exists_other", any_je
        cancelled = f.db.sql("""
            SELECT name FROM `tabJournal Entry`
            WHERE custom_cybervidya_ref LIKE %s AND docstatus=2 LIMIT 1
        """, (f"{ref}__CANCELLED__%",), as_dict=True)
        if cancelled:
            return "previously_cancelled", cancelled[0].name
        return "would_create", None


# ---------------------------------------------------------------------------
# Planning (Phase 3) — build JE plan from parsed sheets + resolution
# ---------------------------------------------------------------------------
def build_plan(sheets, resolver, exclude=None):
    """
    Returns:
      resolved: OrderedDict[sheet] -> {company, abbr, method}
      unresolved: OrderedDict[sheet] -> {reason, rows, min, max, total}
      acct_plan: OrderedDict[sheet] -> {
            'matched': [(nom, account)], 'maprow': [(nom, account)],
            'create': [(nom, planned_name, parent)],
            'review': [(nom, [candidates])],
            'errors': [str], }
      je_plan: list of dict {sheet, company, abbr, date, kind, target, amount, ref}
      head_status: OrderedDict[sheet] -> {recv, cash_needed, cash_ok}
    """
    exclude = set(exclude or [])
    resolved = OrderedDict()
    unresolved = OrderedDict()
    excluded = OrderedDict()
    acct_plan = OrderedDict()
    head_status = OrderedDict()
    je_rows = []

    for sn, data in sheets.items():
        if sn in exclude:
            total = sum((c["amount"] for c in data["cells"]), Decimal(0))
            excluded[sn] = {"rows": data["data_rows"], "min": data["min_date"],
                            "max": data["max_date"], "total": total}
            continue
        company, abbr, method = resolver.resolve_company(sn)
        if not company:
            total = sum((c["amount"] for c in data["cells"]), Decimal(0))
            unresolved[sn] = {
                "reason": method, "rows": data["data_rows"],
                "min": data["min_date"], "max": data["max_date"], "total": total,
            }
            continue
        resolved[sn] = {"company": company, "abbr": abbr, "method": method}

        # --- head verification ---
        recv_name = f"Student Receivable Cybervidya - {abbr}"
        recv_ok, recv_why = resolver.head_exists_leaf(recv_name)
        has_cash = any(c["kind"] == "cash" for c in data["cells"])
        cash_name = f"Cash Cyber Vidhya - {abbr}"
        cash_ok, cash_why = (resolver.head_exists_leaf(cash_name)
                             if has_cash else (True, "n/a"))
        head_status[sn] = {
            "recv_name": recv_name, "recv_ok": recv_ok, "recv_why": recv_why,
            "cash_name": cash_name, "cash_needed": has_cash,
            "cash_ok": cash_ok, "cash_why": cash_why,
        }

        ap = {"matched": [], "maprow": [], "create": [], "review": [], "errors": []}
        if not recv_ok:
            ap["errors"].append(f"Receivable head {recv_name!r} {recv_why} — SHEET SKIPPED")
        if has_cash and not cash_ok:
            ap["errors"].append(f"Cash head {cash_name!r} {cash_why} — SHEET SKIPPED")

        # --- resolve each distinct bank nomenclature ---
        bank_noms = OrderedDict()  # clean_nom -> {acct_no, raw_nom}
        for c in data["cells"]:
            if c["kind"] == "bank":
                bank_noms.setdefault(c["clean_nom"],
                                     {"acct_no": c["acct_no"], "raw_nom": c["raw_nom"]})
        nom_to_ledger = {}
        for cnom, meta in bank_noms.items():
            res = resolver.find_bank_ledger(company, meta["acct_no"], meta["raw_nom"], sn)
            if res["action"] == "matched":
                ap["matched"].append((cnom, res["account"]))
                nom_to_ledger[cnom] = res["account"]
            elif res["action"] == "maprow":
                ap["maprow"].append((cnom, res["account"]))
                nom_to_ledger[cnom] = res["account"]
            elif res["action"] == "create":
                planned = resolver.planned_account_name(company, cnom)
                if not res["parent"]:
                    ap["errors"].append(
                        f"No Bank-Accounts parent group for {company!r}; "
                        f"cannot place {cnom!r} — nomenclature skipped")
                    nom_to_ledger[cnom] = None
                else:
                    ap["create"].append((cnom, planned, res["parent"]))
                    nom_to_ledger[cnom] = planned  # planned name (created in P5)
            elif res["action"] == "review":
                ap["review"].append((cnom, res["candidates"]))
                nom_to_ledger[cnom] = None  # skip until disambiguated
        acct_plan[sn] = ap

        sheet_blocked = bool(not recv_ok or (has_cash and not cash_ok))

        # --- aggregate cells -> JE plan ---
        # key: (date, target_ledger) ; cash target is the cash head
        agg = defaultdict(lambda: Decimal(0))
        meta = {}
        for c in data["cells"]:
            if c["kind"] == "cash":
                target = cash_name
                ref = f"HIST-{sn}-CASH-{ddmy(c['date'])}"
            else:  # bank
                target = nom_to_ledger.get(c["clean_nom"])
                if target is None:
                    continue  # review/skip
                acct = c["acct_no"]
                token = acct[-4:] if acct else slug6(c["clean_nom"])
                ref = f"HIST-{sn}-{token}-{ddmy(c['date'])}"
            key = (c["date"], target, ref)
            agg[key] += c["amount"]
            meta[key] = {"kind": c["kind"]}

        if sheet_blocked:
            continue  # planned but will be skipped; still reported in acct/head

        for (date, target, ref), amount in agg.items():
            je_rows.append({
                "sheet": sn, "company": company, "abbr": abbr,
                "date": date, "kind": meta[(date, target, ref)]["kind"],
                "target": target, "amount": amount, "ref": ref,
                "recv": recv_name,
            })

    return (resolved, unresolved, excluded, acct_plan,
            je_plan_sort(je_rows), head_status)


def je_plan_sort(rows):
    return sorted(rows, key=lambda r: (r["sheet"], r["date"], r["target"]))


# ---------------------------------------------------------------------------
# Dry-run report (Phase 4)
# ---------------------------------------------------------------------------
def money(d):
    return f"{d:,.2f}"


def build_dryrun_report(path, sheets, resolved, unresolved, excluded, acct_plan,
                        je_plan, head_status, resolver, stamp):
    L = []
    def w(s=""):
        L.append(s)

    total_cells = sum(len(s["cells"]) for s in sheets.values())
    all_dates = [d for s in sheets.values() for d in (s["min_date"], s["max_date"]) if d]
    dmin = dstr(min(all_dates)) if all_dates else "—"
    dmax = dstr(max(all_dates)) if all_dates else "—"

    w(f"# Historical CyberVidya Import — DRY RUN")
    w(f"_Generated {stamp} • mode=dryrun • NO database writes performed_")
    w("")
    w("## 1. Source summary")
    w("")
    w(f"- **File:** `{path}`")
    w(f"- **Sheets:** {len(sheets)}")
    w(f"- **Date range (all sheets):** {dmin} → {dmax}")
    w(f"- **Non-zero amount cells parsed (excl. JV-Others):** {total_cells}")
    w("")

    w("## 2. Sheet resolution")
    w("")
    w("| Sheet (CV code) | Resolved company | Abbr | Via | Status |")
    w("|---|---|---|---|---|")
    for sn in sheets:
        if sn in resolved:
            r = resolved[sn]
            w(f"| {sn} | {r['company']} | {r['abbr']} | {r['method']} | RESOLVED |")
        elif sn in excluded:
            w(f"| {sn} | — | — | — | EXCLUDED (--exclude) |")
        else:
            u = unresolved[sn]
            w(f"| {sn} | — | — | — | **UNRESOLVED** ({u['reason']}) |")
    if unresolved:
        w("")
        w("### Unresolved sheets (skipped entirely — no JEs, no accounts)")
        w("")
        w("| Sheet | Data rows | First | Last | Total amount |")
        w("|---|---|---|---|---|")
        for sn, u in unresolved.items():
            w(f"| {sn} | {u['rows']} | {dstr(u['min']) if u['min'] else '—'} "
              f"| {dstr(u['max']) if u['max'] else '—'} | {money(u['total'])} |")
    if excluded:
        w("")
        w("### Excluded sheets (explicitly excluded via --exclude — not imported)")
        w("")
        w("| Sheet | Data rows | First | Last | Total amount |")
        w("|---|---|---|---|---|")
        for sn, u in excluded.items():
            w(f"| {sn} | {u['rows']} | {dstr(u['min']) if u['min'] else '—'} "
              f"| {dstr(u['max']) if u['max'] else '—'} | {money(u['total'])} |")
    w("")

    w("## 3. Account preparation (per resolved company)")
    w("")
    for sn in resolved:
        ap = acct_plan[sn]
        hs = head_status[sn]
        r = resolved[sn]
        w(f"### {sn} → {r['company']} ({r['abbr']})")
        w("")
        recv_flag = "OK" if hs["recv_ok"] else f"**{hs['recv_why'].upper()}**"
        w(f"- Receivable head `{hs['recv_name']}`: {recv_flag}")
        if hs["cash_needed"]:
            cash_flag = "OK" if hs["cash_ok"] else f"**{hs['cash_why'].upper()}**"
            w(f"- Cash head `{hs['cash_name']}`: {cash_flag}")
        else:
            w(f"- Cash head: not needed (no cash columns)")
        if ap["matched"]:
            w(f"- Bank ledgers matched to existing accounts ({len(ap['matched'])}):")
            for nom, acc in ap["matched"]:
                w(f"    - `{nom}` → `{acc}`")
        if ap["maprow"]:
            w(f"- Bank ledgers via existing CyberVidya Bank Map ({len(ap['maprow'])}):")
            for nom, acc in ap["maprow"]:
                w(f"    - `{nom}` → `{acc}`")
        if ap["create"]:
            w(f"- Bank ledgers that WOULD BE CREATED ({len(ap['create'])}):")
            for nom, planned, parent in ap["create"]:
                w(f"    - `{nom}` → new `{planned}`  (under `{parent}`)")
        if ap["review"]:
            w(f"- **MANUAL REVIEW — multiple candidate ledgers ({len(ap['review'])}):**")
            for nom, cands in ap["review"]:
                w(f"    - `{nom}` → candidates: {cands}  (SKIPPED until disambiguated)")
        if ap["errors"]:
            for e in ap["errors"]:
                w(f"- **HARD ERROR:** {e}")
        w("")

    # Phase 5 pre-flight: fiscal year coverage
    w("## 4. Fiscal-year pre-flight (Phase 5 will fail without this)")
    w("")
    w("| Company | Sample date | FY covers? | Fiscal Year |")
    w("|---|---|---|---|")
    checked = {}
    for sn in resolved:
        comp = resolved[sn]["company"]
        if comp in checked:
            continue
        sample = sheets[sn]["min_date"] or (2026, 4, 1)
        ok, fy = resolver.fiscal_year_covers(comp, sample)
        checked[comp] = ok
        w(f"| {comp} | {dstr(sample)} | {'YES' if ok else '**NO**'} | {fy or '—'} |")
    missing_fy = [c for c, ok in checked.items() if not ok]
    if missing_fy:
        w("")
        w(f"> **{len(missing_fy)} company(ies) NOT covered by an active Fiscal Year for "
          f"the import dates.** Phase 5 JE creation will raise FiscalYearError for these. "
          f"Add them to the relevant Fiscal Year's company list before executing:")
        for c in missing_fy:
            w(f"> - {c}")
    w("")

    # JE plan summary
    w("## 5. Journal Entry plan summary")
    w("")
    total_amt = sum((j["amount"] for j in je_plan), Decimal(0))
    n_bank = sum(1 for j in je_plan if j["kind"] == "bank")
    n_cash = sum(1 for j in je_plan if j["kind"] == "cash")
    w(f"- **Total JEs planned:** {len(je_plan)}")
    w(f"- By channel: {n_bank} bank, {n_cash} cash")
    w(f"- **Total amount:** {money(total_amt)}")
    w("")
    w("| Company | JEs | Bank | Cash | Amount |")
    w("|---|---|---|---|---|")
    by_co = defaultdict(lambda: {"n": 0, "bank": 0, "cash": 0, "amt": Decimal(0)})
    for j in je_plan:
        b = by_co[j["company"]]
        b["n"] += 1
        b[j["kind"]] += 1
        b["amt"] += j["amount"]
    for co, b in sorted(by_co.items()):
        w(f"| {co} | {b['n']} | {b['bank']} | {b['cash']} | {money(b['amt'])} |")
    w("")

    # Idempotency precheck
    w("## 6. Idempotency pre-check")
    w("")
    buckets = defaultdict(int)
    samples = defaultdict(list)
    for j in je_plan:
        status, existing = resolver.je_for_ref(j["ref"])
        buckets[status] += 1
        if len(samples[status]) < 5:
            samples[status].append((j["ref"], existing))
    labels = {
        "would_create": "Will CREATE (ref not seen)",
        "already_exists_active": "Will SKIP (active JE already has this ref)",
        "already_exists_other": "Will SKIP (a draft/other JE has this ref)",
        "previously_cancelled": "Will CREATE fresh (only cancelled JEs carry this ref)",
    }
    w("| Outcome | Count |")
    w("|---|---|")
    for k in ("would_create", "previously_cancelled", "already_exists_active", "already_exists_other"):
        if buckets.get(k):
            w(f"| {labels[k]} | {buckets[k]} |")
    w("")

    # JV others
    w("## 7. JV Institute Others — intentionally skipped")
    w("")
    w("| Sheet | JV rows | JV total amount |")
    w("|---|---|---|")
    for sn, s in sheets.items():
        jv = s["skipped_jv"]
        if jv["rows"] or jv["amount"]:
            w(f"| {sn} | {jv['rows']} | {money(jv['amount'])} |")
    w("")

    # Empty-nomenclature columns ignored
    empties = [(sn, s["empty_cols"]) for sn, s in sheets.items() if s["empty_cols"]]
    if empties:
        w("### Empty-nomenclature columns ignored (no row-2 label)")
        w("")
        for sn, cols in empties:
            w(f"- {sn}: columns {cols}")
        w("")

    # Top of JE list
    w("## 8. First 20 planned JEs (sanity sample)")
    w("")
    w("| Sheet | Date | Channel | Target ledger | Amount | Ref |")
    w("|---|---|---|---|---|---|")
    for j in je_plan[:20]:
        w(f"| {j['sheet']} | {dstr(j['date'])} | {j['kind']} | {j['target']} "
          f"| {money(j['amount'])} | {j['ref']} |")
    w("")
    w("---")
    w("")
    w("**Dry run complete. Awaiting \"Go\" to execute Phase 5 against erp.jewonline.in.**")

    return "\n".join(L)


# ---------------------------------------------------------------------------
# Phase 5 — execute (guarded)
# ---------------------------------------------------------------------------
def execute(frappe, sheets, resolved, acct_plan, je_plan, head_status, resolver,
            out_dir, stamp):
    frappe.set_user("Administrator")
    created_accounts = []
    created_map_parents = []
    created_map_rows = []
    errors = []

    # --- 5a: create accounts + mapping rows first ---
    for sn in resolved:
        company = resolved[sn]["company"]
        abbr = resolved[sn]["abbr"]
        ap = acct_plan[sn]
        if ap["errors"]:
            continue  # blocked sheet
        for cnom, planned, parent in ap["create"]:
            try:
                if not frappe.db.exists("Account", planned):
                    acc = frappe.new_doc("Account")
                    # account_name is the part before ' - {ABBR}'; ERPNext re-suffixes
                    acc.account_name = (planned[:-(len(abbr) + 3)]
                                        if planned.endswith(f" - {abbr}") else planned)
                    acc.company = company
                    acc.parent_account = parent
                    acc.account_type = "Bank"
                    acc.is_group = 0
                    acc.account_currency = frappe.db.get_value("Company", company, "default_currency")
                    acc.insert(ignore_permissions=True)
                    created_accounts.append(acc.name)
                # ensure mapping parent
                if not frappe.db.exists("CyberVidya Account Mapping", sn):
                    m = frappe.new_doc("CyberVidya Account Mapping")
                    m.cybervidya_institution = sn
                    m.company = company
                    m.insert(ignore_permissions=True)
                    created_map_parents.append(sn)
                # find verbatim raw nomenclature for this clean nom
                raw = next((c["raw_nom"] for c in sheets[sn]["cells"]
                            if c["kind"] == "bank" and c["clean_nom"] == cnom), cnom)
                exists_row = frappe.db.sql("""
                    SELECT 1 FROM `tabCyberVidya Bank Map`
                    WHERE parent=%s AND cybervidya_bank=%s LIMIT 1
                """, (sn, raw))
                if not exists_row:
                    parent_doc = frappe.get_doc("CyberVidya Account Mapping", sn)
                    parent_doc.append("bank_accounts", {
                        "cybervidya_bank": raw,
                        "bank_account": planned,
                    })
                    parent_doc.save(ignore_permissions=True)
                    created_map_rows.append((sn, raw, planned))
                frappe.db.commit()
            except Exception as e:
                frappe.db.rollback()
                errors.append({"phase": "account", "sheet": sn, "nom": cnom,
                               "error": f"{type(e).__name__}: {e}",
                               "tb": traceback.format_exc()})

    # --- 5b: create JEs ---
    created, skipped, failed = [], [], []
    n = 0
    for j in je_plan:
        n += 1
        ref = j["ref"]
        try:
            status, existing = resolver.je_for_ref(ref)
            if status in ("already_exists_active", "already_exists_other"):
                skipped.append((ref, existing))
                continue
            je = frappe.new_doc("Journal Entry")
            je.voucher_type = "Journal Entry"
            je.company = j["company"]
            je.posting_date = frappe.utils.getdate(dstr(j["date"]))
            je.user_remark = f"Historical CyberVidya import — {j['kind']} collection {ref}"
            je.custom_cybervidya_ref = ref
            amt = float(j["amount"])
            je.append("accounts", {
                "account": j["target"],
                "debit_in_account_currency": amt,
                "credit_in_account_currency": 0,
            })
            je.append("accounts", {
                "account": j["recv"],
                "debit_in_account_currency": 0,
                "credit_in_account_currency": amt,
            })
            je.insert(ignore_permissions=True)
            je.submit()
            frappe.db.commit()
            created.append((ref, je.name))
        except (frappe.UniqueValidationError, frappe.DuplicateEntryError):
            frappe.db.rollback()
            existing = frappe.db.get_value("Journal Entry",
                                           {"custom_cybervidya_ref": ref}, "name")
            skipped.append((ref, existing))
        except Exception as e:
            frappe.db.rollback()
            failed.append({"ref": ref, "error": f"{type(e).__name__}: {e}",
                           "tb": traceback.format_exc()})
        if n % 25 == 0:
            print(f"  ... {n}/{len(je_plan)} processed "
                  f"({len(created)} created, {len(skipped)} skipped, {len(failed)} failed)")

    # --- report ---
    L = [f"# Historical CyberVidya Import — EXECUTION RUN",
         f"_Generated {stamp} • mode=execute_", ""]
    L.append(f"- JEs created: **{len(created)}**")
    L.append(f"- JEs skipped (already existed): {len(skipped)}")
    L.append(f"- JEs failed: {len(failed)}")
    L.append(f"- Bank ledgers created: {len(created_accounts)}")
    L.append(f"- Mapping parents created: {len(created_map_parents)}")
    L.append(f"- Mapping bank-rows added: {len(created_map_rows)}")
    L.append("")
    if created_accounts:
        L.append("## Newly created bank ledgers")
        for a in created_accounts:
            L.append(f"- {a}")
        L.append("")
    if created_map_rows:
        L.append("## Newly created CyberVidya Bank Map rows")
        for sn, raw, acc in created_map_rows:
            L.append(f"- {sn}: `{raw}` → `{acc}`")
        L.append("")
    # per-company totals
    by_co = defaultdict(lambda: Decimal(0))
    name_to_amt = {j["ref"]: j["amount"] for j in je_plan}
    for ref, _ in created:
        # find company
        j = next((x for x in je_plan if x["ref"] == ref), None)
        if j:
            by_co[j["company"]] += j["amount"]
    L.append("## Posted totals per company (created JEs only)")
    L.append("")
    L.append("| Company | Amount posted |")
    L.append("|---|---|")
    for co, amt in sorted(by_co.items()):
        L.append(f"| {co} | {money(amt)} |")
    L.append("")
    if failed:
        L.append("## Failures")
        for fobj in failed:
            L.append(f"- `{fobj['ref']}`: {fobj['error']}")
        L.append("")

    report = "\n".join(L)
    rpath = os.path.join(out_dir, f"historical_import_run_{stamp}.md")
    with open(rpath, "w") as fh:
        fh.write(report)
    print(report)
    print(f"\n[run report written to {rpath}]")

    if errors or failed:
        epath = os.path.join(out_dir, f"historical_import_errors_{stamp}.log")
        with open(epath, "w") as fh:
            for e in errors:
                fh.write(f"=== {e.get('phase')} {e.get('sheet')} {e.get('nom')} ===\n")
                fh.write(e["tb"] + "\n")
            for fobj in failed:
                fh.write(f"=== JE {fobj['ref']} ===\n{fobj['tb']}\n")
        print(f"[error log written to {epath}]")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--site", required=True)
    ap.add_argument("--file", required=True)
    ap.add_argument("--mode", choices=["dryrun", "execute"], default="dryrun")
    ap.add_argument("--out-dir", default="/tmp")
    ap.add_argument("--stamp", default="run")
    ap.add_argument("--exclude", default="",
                    help="comma-separated sheet names to skip entirely")
    args = ap.parse_args()
    exclude = [s.strip() for s in args.exclude.split(",") if s.strip()]

    if not os.path.exists(args.file):
        print(f"ERROR: file not found: {args.file}", file=sys.stderr)
        sys.exit(2)

    print(f"[reading workbook] {args.file}")
    sheets = parse_workbook(args.file)
    print(f"[parsed] {len(sheets)} sheets")

    os.chdir("/home/frappe/frappe-bench/sites")
    import frappe
    frappe.init(site=args.site)
    frappe.connect()
    try:
        resolver = Resolver(frappe)
        if exclude:
            print(f"[exclude] skipping sheets: {exclude}")
        (resolved, unresolved, excluded, acct_plan,
         je_plan, head_status) = build_plan(sheets, resolver, exclude=exclude)

        if args.mode == "dryrun":
            report = build_dryrun_report(
                args.file, sheets, resolved, unresolved, excluded, acct_plan,
                je_plan, head_status, resolver, args.stamp)
            rpath = os.path.join(args.out_dir, f"historical_import_dryrun_{args.stamp}.md")
            with open(rpath, "w") as fh:
                fh.write(report)
            print(report)
            print(f"\n[dry-run report written to {rpath}]")
        else:
            print("[MODE=execute] creating accounts, mappings, and JEs...")
            execute(frappe, sheets, resolved, acct_plan, je_plan, head_status,
                    resolver, args.out_dir, args.stamp)
    finally:
        frappe.destroy()


if __name__ == "__main__":
    main()
